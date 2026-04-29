"""Parse WeChat XLSX export files."""
import re
import io
from datetime import datetime

from openpyxl import load_workbook


COLUMN_MAP = {
    "交易时间": "transaction_time",
    "交易类型": "transaction_type",
    "交易对方": "counterparty",
    "商品": "product_desc",
    "收/支": "direction_raw",
    "金额(元)": "amount_raw",
    "支付方式": "payment_method",
    "当前状态": "status",
    "交易单号": "source_txn_id",
    "商户单号": "merchant_order",
    "备注": "notes",
}

DIRECTION_MAP = {
    "收入": "income",
    "支出": "expense",
}


def parse_wechat_xlsx(file_bytes: bytes) -> dict:
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    except Exception as e:
        raise ValueError(
            f"无法读取微信账单文件: {e}。"
            "请确认：1) 文件是微信导出的 .xlsx 格式；2) 文件未损坏；3) 文件未设置密码。"
        )

    try:
        ws = wb.active
    except Exception:
        wb.close()
        raise ValueError("无法读取工作表，文件可能已损坏。")

    header_row = 0
    col_map = {}
    date_start, date_end = None, None

    # Scan first 30 rows for header and date range
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, max_col=16, values_only=True), start=1):
        line = " ".join(str(c) for c in row if c)
        m = re.search(r"起始时间.*?\[([\d-]+).*?\].*?终止时间.*?\[([\d-]+).*?\]", line)
        if m:
            date_start = m.group(1)
            date_end = m.group(2)

        if row[0] and str(row[0]) == "交易时间":
            header_row = row_idx
            for col_idx, cell_val in enumerate(row):
                cn_name = str(cell_val).strip() if cell_val else ""
                if cn_name in COLUMN_MAP:
                    col_map[col_idx] = COLUMN_MAP[cn_name]
            break

    if header_row == 0:
        wb.close()
        raise ValueError(
            "未找到表头行（包含'交易时间'），请确认这是微信导出的账单文件。"
            "在微信中：我 → 服务 → 钱包 → 账单 → 导出账单。"
        )

    transactions = []
    skipped = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or all(c is None for c in row):
            continue

        data = {}
        for col_idx, key in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            data[key] = val

        ts_val = data.get("transaction_time")
        if ts_val is None:
            skipped += 1
            continue
        if isinstance(ts_val, datetime):
            ts = ts_val
        else:
            ts_str = str(ts_val).strip()
            try:
                ts = datetime.strptime(ts_str[:19], "%Y-%m-%d %H:%M:%S")
            except ValueError:
                skipped += 1
                continue

        dir_raw = str(data.get("direction_raw", "")).strip()
        direction = DIRECTION_MAP.get(dir_raw)
        if direction is None:
            skipped += 1
            continue

        amount = data.get("amount_raw")
        if amount is None:
            skipped += 1
            continue
        try:
            amount = abs(float(amount))
        except (ValueError, TypeError):
            skipped += 1
            continue

        st = data.get("status")
        status = str(st).strip() if st else ""

        transactions.append({
            "transaction_time": ts,
            "source_platform": "wechat",
            "transaction_type": str(data.get("transaction_type", "")).strip(),
            "counterparty": str(data.get("counterparty", "")).strip(),
            "product_desc": str(data.get("product_desc", "")).strip(),
            "direction": direction,
            "amount": amount,
            "payment_method": str(data.get("payment_method", "")).strip(),
            "status": status,
            "source_txn_id": str(data.get("source_txn_id", "")).strip(),
            "merchant_order": str(data.get("merchant_order", "")).strip(),
            "notes": str(data.get("notes", "")).strip(),
        })

    wb.close()

    if not transactions:
        raise ValueError(
            f"未能提取到有效交易记录（跳过 {skipped} 行）。请确认文件中包含交易数据。"
        )

    return {
        "source_platform": "wechat",
        "transactions": transactions,
        "record_count": len(transactions),
        "date_start": date_start or "",
        "date_end": date_end or "",
    }
